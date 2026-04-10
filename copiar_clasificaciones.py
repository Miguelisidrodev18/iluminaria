import openpyxl

ORIGEN = "Productos.xlsx"
DESTINO = "plantilla_importacion_kyrios (14).xlsx"
HOJA_ORIGEN = "productos_trabajo"
HOJA_NUEVA = "CLASIFICACIONES_CLIENTE"

# Columnas AT:II en índices 1-based
# AT = columna 46, II = columna 243
COL_INICIO = 46   # AT
COL_FIN = 243     # II

print(f"Abriendo {ORIGEN} en modo solo lectura...")
wb_origen = openpyxl.load_workbook(ORIGEN, read_only=True, data_only=True)
ws_origen = wb_origen[HOJA_ORIGEN]

print(f"Leyendo rango AT:II de '{HOJA_ORIGEN}'...")
datos = []
for fila in ws_origen.iter_rows(
    min_col=COL_INICIO,
    max_col=COL_FIN,
    values_only=True
):
    datos.append(fila)

wb_origen.close()

filas_leidas = len(datos)
cols_leidas = len(datos[0]) if datos else 0
print(f"  -> {filas_leidas} filas x {cols_leidas} columnas leídas")

print(f"\nAbriendo {DESTINO}...")
wb_destino = openpyxl.load_workbook(DESTINO)

if HOJA_NUEVA in wb_destino.sheetnames:
    print(f"  Hoja '{HOJA_NUEVA}' ya existe, eliminando para recrear...")
    del wb_destino[HOJA_NUEVA]

ws_nueva = wb_destino.create_sheet(title=HOJA_NUEVA)
print(f"  Hoja '{HOJA_NUEVA}' creada al final del workbook.")

print(f"Escribiendo datos en '{HOJA_NUEVA}'...")
for i, fila in enumerate(datos, start=1):
    for j, valor in enumerate(fila, start=1):
        if valor is not None:
            ws_nueva.cell(row=i, column=j, value=valor)

print(f"Guardando {DESTINO}...")
wb_destino.save(DESTINO)
print(f"\nListo. Se copiaron {filas_leidas} filas y {cols_leidas} columnas a la hoja '{HOJA_NUEVA}'.")
