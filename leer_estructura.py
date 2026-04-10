import openpyxl

wb = openpyxl.load_workbook("plantilla_importacion_kyrios (14).xlsx", read_only=True, data_only=True)

print("=== HOJAS ===")
for nombre in wb.sheetnames:
    print(f"  - {nombre}")

print()

for nombre in wb.sheetnames:
    ws = wb[nombre]
    print(f"\n{'='*60}")
    print(f"HOJA: {nombre}")
    print(f"{'='*60}")
    filas_con_dato = 0
    for i, fila in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(v is not None for v in fila):
            filas_con_dato += 1
            valores = [str(v) if v is not None else '' for v in fila]
            # Truncar valores largos
            valores = [v[:30] if len(v) > 30 else v for v in valores]
            print(f"  Fila {i:3d}: {valores}")
            if filas_con_dato >= 8:
                print(f"  ... (mostrando solo primeras 8 filas con datos)")
                break

wb.close()
